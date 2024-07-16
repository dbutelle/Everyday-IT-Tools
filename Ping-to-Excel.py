#### DODOW62 ####
#### Version 1.0 ####


import subprocess
import pandas as pd
import os
from openpyxl import Workbook, load_workbook  # Import load_workbook from openpyxl
from openpyxl.styles import PatternFill

def ping_ip(ip):
    """Perform a ping on a single IP address and return True if it responds, False otherwise."""
    try:
        result = subprocess.run(['ping', '-n', '1', ip], capture_output=True, text=True, timeout=2)
        if "TTL=" in result.stdout:
            return True
        else:
            return False
    except subprocess.TimeoutExpired:
        return False

def ping_multiple_ips(ip_list):
    """Perform ping on a list of IP addresses and return a list of dictionaries with IP, status, and color."""
    results = []
    for ip in ip_list:
        if ping_ip(ip):
            results.append({'IP': ip, 'Status': 'Responded'})
        else:
            results.append({'IP': ip, 'Status': 'No response'})
    return results

def main():
    current_dir = os.path.dirname(__file__)
    ip_list = []
    
    print("Enter IP addresses to ping (enter 'quit' or 'exit' to finish):")
    while True:
        ip = input("IP address: ")
        if ip.lower() == 'quit' or ip.lower() == 'exit':
            break
        ip_list.append(ip)
    
    if ip_list:
        # Perform ping on the list of IP addresses
        ping_results = ping_multiple_ips(ip_list)
        
        # Create DataFrame
        df = pd.DataFrame(ping_results)
        
        # Export to Excel with conditional formatting
        output_file = os.path.join(current_dir, 'ping_results.xlsx')
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Apply conditional formatting
        wb = load_workbook(output_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, max_row=len(ip_list) + 1, min_col=2, max_col=2):
            for cell in row:
                if cell.value == 'Responded':
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif cell.value == 'No response':
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        wb.save(output_file)
        print(f"Ping results exported to {output_file}")
    else:
        print("No IP addresses entered. Exiting...")

if __name__ == "__main__":
    main()

